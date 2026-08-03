"""Importador del catálogo espacial desde ReporteUbicaciones.xlsx.

    python tools/import_spatial_catalog.py <ruta.xlsx> --warehouse-code WH-001
    python tools/import_spatial_catalog.py <ruta.xlsx> --warehouse-code WH-001 --dry-run

QUÉ CREA, y por qué esos números:

    1 sitio          `DEFAULT`, sin validar. El catálogo NO dice a qué sitio físico
                     pertenece cada ubicación: `Preámbulo` es ortogonal a
                     `IdSucursal` (medido), así que no se afirma nada.
    347 nodos        uno por `Referencia` distinta, en correspondencia 1:1 con
                     `IdAlmacenamiento`.
    2.701 cuerpos    uno por par (Referencia, Columna) distinto.
    29.310 huecos    uno por fila. NUNCA se genera la posición hermana: 3.866
                     tripletas (ref, col, nivel) tienen UNA sola posición.

QUÉ NO CREA
    · pasillos — `aisle` queda vacío. La familia de letras del código abarca 2
      preámbulos, 2 tipos y 11 zonas: no es un pasillo. Los racks quedan como
      nodos RAÍZ.
    · posiciones hermanas · world_* · nodos de nivel · nodos de posición.

IDEMPOTENCIA, en tres niveles:
    1. `sha256` del archivo, registrado en el lote. Reimportar el mismo archivo se
       detecta antes de leer una fila.
    2. `ON CONFLICT` sobre la clave natural de cada entidad.
    3. El lote es una transacción: o entra entero o no entra nada.

TOLERANCIA A FILAS INVÁLIDAS: una fila mala no tumba el lote. Se registra con su
número, su contenido crudo y el motivo, y el resumen las cuenta por motivo.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

import asyncpg
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

# Reutiliza la resolución de conexión de admin_sql: host y base de `.env.local`,
# contraseña de postgres de `docs/.envlocal`. Nunca imprime la contraseña.
from admin_sql import _connection_kwargs  # noqa: E402

# ── Encabezados esperados, por nombre Y posición ────────────────────────────
# Un archivo con columnas reordenadas se rechaza entero: durante el análisis un
# índice desplazado un puesto produjo resultados plausibles y falsos.
ENCABEZADOS = [
    "Id Almacenamiento", "Ubicación", "Preámbulo", "Referencia", "Columna",
    "Nivel", "Posición", "Tipo Ubicación", "Estado", "Situación",
    "Eje X", "Eje Y", "Eje Z", "Peso Máximo", "Zona Almacenaje",
    "Zona Picking", "Zona Trabajo Recurso", "Zona Trabajo Preparación",
    "Zona Cola Preparación", "Id Ubicación",
]

# Techo de plausibilidad de capacidad, en kg. Debe coincidir con
# `core.capacity_ceiling('weight_kg')`; se COMPRUEBA contra la base antes de
# escribir, porque dos umbrales copiados divergen en cuanto uno se toca.
#
# No es una lista de centinelas: el catálogo real usa seis grafías distintas de
# «sin límite» (1e5, 1e6, 9999999, 1e7, 99999999, 1e8) y enumerarlas fue el
# defecto que 0058 corrigió. Ver la cabecera de 0058 para los datos medidos.
TECHO_PESO_KG = 50000

# Estado del WMS → estado del ESPACIO. `OCUP` no entra: la ocupación es del
# snapshot de `wms`, no del estante (SPA-11 y SPA-12).
ESTADO_WMS = {"DISP": "available", "BLOQ": "blocked", "OCUP": "available"}

PATRON_CODIGO = re.compile(r"^[A-Z0-9][A-Z0-9._]*-C[0-9]{3}-N[0-9]{2}-[0-9]$")

_ACENTOS = str.maketrans("ÁÀÄÂÃÅÉÈËÊÍÌÏÎÓÒÖÔÕÚÙÜÛÑÇ ", "AAAAAAEEEEIIIIOOOOOUUUUNC_")


def normalizar(bruto: str | None) -> str:
    """Misma regla que `core.normalize_spatial_code()`, y tiene que coincidir.

    El espacio pasa a `_`, NUNCA a `-`: el guion separa segmentos, y `PHA LO`
    convertido a `PHA-LO` daría cinco segmentos.
    """
    if bruto is None:
        return ""
    limpio = unicodedata.normalize("NFC", str(bruto).strip().upper()).translate(_ACENTOS)
    return re.sub(r"[^A-Z0-9._-]", "_", limpio)


def texto(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def entero(v: Any) -> int | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None


def capacidad(v: Any) -> tuple[int | None, int | None]:
    """Devuelve (capacidad utilizable, valor crudo si se descartó).

    Un valor por encima del techo no es una capacidad, es «sin límite» escrito
    con un número. Se anula, pero se DEVUELVE también el crudo: el valor no
    dice nada de la ubicación y sí dice algo del WMS de origen.
    """
    n = entero(v)
    if n is None or n == 0:
        return None, None
    return (None, n) if n >= TECHO_PESO_KG else (n, None)


class RechazoError(Exception):
    def __init__(self, motivo: str, campo: str = "") -> None:
        super().__init__(motivo)
        self.motivo = motivo
        self.campo = campo


def sha256_de(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return h.hexdigest()


def leer_catalogo(ruta: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Devuelve (filas válidas, rechazos). No tumba el lote por una fila mala."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)

    cabecera = [texto(c) for c in next(it)]
    if cabecera[: len(ENCABEZADOS)] != ENCABEZADOS:
        wb.close()
        faltan = [h for h in ENCABEZADOS if h not in (cabecera or [])]
        raise SystemExit(
            f"ENCABEZADOS INESPERADOS. Se esperaban {len(ENCABEZADOS)} en este orden.\n"
            f"  faltan o cambiaron de sitio: {faltan or 'ninguno, pero el orden difiere'}\n"
            f"  recibido: {cabecera[:6]}…"
        )
    idx = {c: i for i, c in enumerate(cabecera)}

    filas: list[dict[str, Any]] = []
    rechazos: list[dict[str, Any]] = []

    for n, cruda in enumerate(it, start=2):
        try:
            ref_bruta = texto(cruda[idx["Referencia"]])
            ubic_bruta = texto(cruda[idx["Ubicación"]])
            if not ref_bruta:
                raise RechazoError("Referencia vacía", "Referencia")
            if not ubic_bruta:
                raise RechazoError("Ubicación vacía", "Ubicación")

            col = entero(cruda[idx["Columna"]])
            niv = entero(cruda[idx["Nivel"]])
            pos = entero(cruda[idx["Posición"]])
            if col is None or niv is None or pos is None:
                raise RechazoError("Columna, Nivel o Posición ausente", "Columna/Nivel/Posición")
            if not (1 <= niv <= 99):
                raise RechazoError(f"Nivel {niv} fuera de 1..99", "Nivel")
            if not (1 <= pos <= 9):
                raise RechazoError(f"Posición {pos} fuera de 1..9", "Posición")
            if col < 1:
                raise RechazoError(f"Columna {col} inválida", "Columna")

            ref_norm = normalizar(ref_bruta)
            code = f"{ref_norm}-C{col:03d}-N{niv:02d}-{pos}"

            # La clasificación no puede mentir: si no cumple el patrón, es `opaque`
            # y el parser estructurado NO se le aplica.
            forma = "structured" if PATRON_CODIGO.match(code) else "opaque"

            peso_max, peso_crudo = capacidad(cruda[idx["Peso Máximo"]])

            filas.append(
                {
                    "fila": n,
                    "ref_bruta": ref_bruta,
                    "ref_norm": ref_norm,
                    "storage_id": texto(cruda[idx["Id Almacenamiento"]]),
                    "preambulo": texto(cruda[idx["Preámbulo"]]),
                    "col": col, "niv": niv, "pos": pos,
                    "code": code,
                    "external_code": ubic_bruta,
                    "external_location_id": texto(cruda[idx["Id Ubicación"]]),
                    "tipo_wms": texto(cruda[idx["Tipo Ubicación"]]),
                    "estado_wms": texto(cruda[idx["Estado"]]),
                    "situacion": texto(cruda[idx["Situación"]]),
                    "lx": entero(cruda[idx["Eje X"]]),
                    "ly": entero(cruda[idx["Eje Y"]]),
                    "lz": entero(cruda[idx["Eje Z"]]),
                    "peso_max": peso_max,
                    "peso_max_crudo": peso_crudo,
                    "zona": texto(cruda[idx["Zona Almacenaje"]]),
                    "forma": forma,
                    "crudo": {
                        k: (str(cruda[i]) if cruda[i] is not None else None)
                        for k, i in idx.items() if k
                    },
                }
            )
        except RechazoError as r:
            rechazos.append(
                {"fila": n, "motivo": r.motivo, "campo": r.campo,
                 "crudo": json.dumps([str(c) if c is not None else None for c in cruda[:8]])}
            )

    wb.close()
    return filas, rechazos


async def importar(ruta: Path, wh_code: str, *, dry_run: bool, force: bool = False) -> int:
    sha = sha256_de(ruta)
    print(f"→ archivo : {ruta.name}")
    print(f"  sha256  : {sha}")
    print(f"  almacen : {wh_code}")

    filas, rechazos = leer_catalogo(ruta)
    print(f"  filas leidas : {len(filas):,}   rechazadas: {len(rechazos)}")
    if rechazos:
        por_motivo = Counter(r["motivo"] for r in rechazos)
        for m, c in por_motivo.most_common():
            print(f"     {c:5d} x {m}")

    # ── Agregados: qué nodos hay que crear ──────────────────────────────────
    refs: dict[str, dict[str, Any]] = {}
    for f in filas:
        r = refs.setdefault(
            f["ref_norm"],
            {"externo": f["ref_bruta"], "storage_id": f["storage_id"],
             "preambulo": f["preambulo"], "cols": set(), "tipos": set(), "zonas": set()},
        )
        r["cols"].add(f["col"])
        if f["tipo_wms"]:
            r["tipos"].add(f["tipo_wms"])
        if f["zona"]:
            r["zonas"].add(f["zona"])

    bays = {(f["ref_norm"], f["col"]) for f in filas}
    print(f"  nodos principales : {len(refs):,}")
    print(f"  cuerpos           : {len(bays):,}")
    print(f"  ubicaciones       : {len(filas):,}")

    if dry_run:
        print("  DRY-RUN: no se escribe nada")
        return 0

    kw = _connection_kwargs()
    print(f"→ {kw['user']}@{kw['host']}:{kw['port']}/{kw['database']}")
    conn = await asyncpg.connect(**kw)
    try:
        async with conn.transaction():
            wh = await conn.fetchrow(
                "SELECT id, tenant_id, name FROM core.warehouses "
                "WHERE code = $1 AND deleted_at IS NULL", wh_code)
            if wh is None:
                raise SystemExit(f"No existe el almacen {wh_code}")
            tid, wid = wh["tenant_id"], wh["id"]
            print(f"  almacen resuelto: {wh['name']}")

            # El techo del importador y el del motor tienen que ser EL MISMO
            # número. Comprobarlo cuesta una consulta; no comprobarlo produce un
            # importador que anula capacidades que la base habría aceptado, o —
            # peor — que envía valores que la base rechaza a mitad del lote.
            techo_bd = await conn.fetchval("SELECT core.capacity_ceiling('weight_kg')")
            if techo_bd is None or int(techo_bd) != TECHO_PESO_KG:
                raise SystemExit(
                    f"El techo de capacidad del importador ({TECHO_PESO_KG}) no coincide "
                    f"con core.capacity_ceiling('weight_kg') = {techo_bd}. "
                    "Aline los dos antes de importar."
                )

            # ── Idempotencia nivel 1 · el hash del archivo ──────────────────
            ya = await conn.fetchval(
                "SELECT count(1) FROM spatial.import_batches "
                "WHERE tenant_id=$1 AND warehouse_id=$2 AND file_sha256=$3 AND status='completed'",
                tid, wid, sha)
            if ya and not force:
                print(f"  Este archivo ya se importo (sha256 {sha[:12]}…). Nada que hacer.")
                print("  (--force reejecuta los upserts; sirve para PROBAR la idempotencia)")
                return 0
            if ya:
                # El índice único de lotes es PARCIAL sobre `completed`, así que un
                # segundo lote completado con el mismo sha lo violaría. Se marca el
                # anterior como reemplazado por este.
                await conn.execute(
                    "UPDATE spatial.import_batches SET status='failed', "
                    "  finished_at = coalesce(finished_at, now()) "
                    "WHERE tenant_id=$1 AND warehouse_id=$2 AND file_sha256=$3 "
                    "  AND status='completed'", tid, wid, sha)
                print("  --force: se reejecutan los upserts sobre datos ya presentes")

            lote = await conn.fetchval(
                "INSERT INTO spatial.import_batches "
                "(tenant_id, warehouse_id, source_name, file_sha256, rows_read, "
                " rows_rejected, status) "
                "VALUES ($1,$2,$3,$4,$5,$6,'running') RETURNING id",
                tid, wid, ruta.name, sha, len(filas) + len(rechazos), len(rechazos))

            for r in rechazos:
                await conn.execute(
                    "INSERT INTO spatial.import_row_errors "
                    "(batch_id, tenant_id, row_number, reason_code, field_name, raw_row) "
                    "VALUES ($1,$2,$3,$4,$5,$6)",
                    lote, tid, r["fila"], r["motivo"][:60], r["campo"][:40], r["crudo"])

            # ── Sitio único, sin validar ────────────────────────────────────
            site = await conn.fetchval(
                "INSERT INTO spatial.sites (tenant_id, warehouse_id, name, code, "
                "  is_validated, raw_source) "
                "VALUES ($1,$2,'Sitio unico (sin validar)','DEFAULT',false,$3) "
                "ON CONFLICT (tenant_id, warehouse_id, code) WHERE deleted_at IS NULL "
                "  DO UPDATE SET updated_at = now() "
                "RETURNING id",
                tid, wid, json.dumps({"created_by": "import_spatial_catalog"}))

            # ── 347 nodos principales · RAÍZ, sin pasillo ───────────────────
            #
            # ⚠ INSERCIÓN DE CONJUNTO, no fila a fila. La primera versión hacía un
            #   `fetchval` por nodo: medido, 275 ms por fila contra el pooler de AWS
            #   —14 minutos sin terminar los 3.048 nodos— porque cada fila era un
            #   viaje de red completo. Con `unnest` son DOS sentencias en total.
            orden = sorted(refs.items())
            await conn.execute(
                "INSERT INTO spatial.nodes "
                "(tenant_id, warehouse_id, site_id, parent_node_id, node_type, "
                " node_function, node_code, external_code, name, logical_index, "
                " external_site_code, external_storage_id, raw_source) "
                "SELECT $1, $2, $3, NULL, 'rack', nf.code, t.code, t.externo, t.externo, "
                "       t.indice, t.preambulo, t.storage_id, t.crudo::jsonb "
                "  FROM unnest($4::text[], $5::text[], $6::int[], $7::text[], $8::text[], "
                "              $9::text[], $10::text[]) "
                "         AS t(code, externo, indice, preambulo, storage_id, tipo_wms, crudo) "
                "  LEFT JOIN spatial.node_functions nf ON nf.wms_type_code = t.tipo_wms "
                "ON CONFLICT (tenant_id, warehouse_id, node_code) WHERE deleted_at IS NULL "
                "DO UPDATE SET external_code = EXCLUDED.external_code, "
                "              logical_index = EXCLUDED.logical_index, "
                "              node_function = EXCLUDED.node_function, "
                "              raw_source = EXCLUDED.raw_source, updated_at = now()",
                tid, wid, site,
                [n for n, _ in orden],
                [i["externo"] for _, i in orden],
                [int(m.group(1)) if (m := re.search(r"(\d+)", n)) else None for n, _ in orden],
                [i["preambulo"] for _, i in orden],
                [i["storage_id"] for _, i in orden],
                [sorted(i["tipos"])[0] if i["tipos"] else None for _, i in orden],
                [json.dumps({"source": "ReporteUbicaciones", "zonas": sorted(i["zonas"]),
                             "tipos_wms": sorted(i["tipos"]), "columnas": len(i["cols"])})
                 for _, i in orden],
            )
            mapa_ref = {
                r["node_code"]: r["id"] for r in await conn.fetch(
                    "SELECT id, node_code FROM spatial.nodes "
                    "WHERE warehouse_id = $1 AND node_type = 'rack' AND deleted_at IS NULL", wid)
            }
            print(f"  nodos principales : {len(mapa_ref):,}")

            # ── 2.701 cuerpos · una sola sentencia, resolviendo el padre por JOIN ──
            pares = sorted(bays)
            await conn.execute(
                "INSERT INTO spatial.nodes "
                "(tenant_id, warehouse_id, site_id, parent_node_id, node_type, "
                " node_code, name, logical_index) "
                "SELECT $1, $2, $3, r.id, 'bay', "
                "       t.ref || '-C' || lpad(t.col::text, 3, '0'), "
                "       'Cuerpo C' || lpad(t.col::text, 3, '0') || ' de ' || t.ref, t.col "
                "  FROM unnest($4::text[], $5::int[]) AS t(ref, col) "
                "  JOIN spatial.nodes r ON r.warehouse_id = $2 AND r.node_code = t.ref "
                "                      AND r.node_type = 'rack' AND r.deleted_at IS NULL "
                "ON CONFLICT (tenant_id, warehouse_id, node_code) WHERE deleted_at IS NULL "
                "DO UPDATE SET logical_index = EXCLUDED.logical_index, updated_at = now()",
                tid, wid, site, [r for r, _ in pares], [c for _, c in pares],
            )
            # Clave (id_del_padre, índice) en lugar de partir `node_code`: el código
            # normalizado admite guiones, así que un `rsplit('-C')` sería frágil.
            mapa_bay = {
                (r["parent_node_id"], r["logical_index"]): r["id"]
                for r in await conn.fetch(
                    "SELECT id, parent_node_id, logical_index FROM spatial.nodes "
                    "WHERE warehouse_id = $1 AND node_type = 'bay' AND deleted_at IS NULL", wid)
            }
            print(f"  cuerpos           : {len(mapa_bay):,}")

            # ── 29.310 ubicaciones · inserción de conjunto, en lotes ─────────
            #
            # Un `executemany` con 29.310 filas sigue siendo 29.310 ejecuciones del
            # plan, cada una con su ida y vuelta de bind/execute. `unnest` de columnas
            # paralelas lo convierte en UNA sentencia por lote: 15 viajes en total.
            sql_loc = (
                "INSERT INTO spatial.locations "
                "(tenant_id, warehouse_id, node_id, code, external_code, "
                " external_location_id, type, code_form, logical_column, "
                " logical_level, logical_position, logical_x, logical_y, logical_z, "
                " max_weight_kg, status, location_situation, origin, raw_source) "
                "SELECT $1, $2, u.node_id, u.code, u.ext_code, u.ext_id, 'rack', u.forma, "
                "       u.col, u.niv, u.pos, u.lx, u.ly, u.lz, u.peso, u.estado, "
                "       u.situacion, 'catalog', u.crudo::jsonb "
                "  FROM unnest($3::uuid[], $4::text[], $5::text[], $6::text[], $7::text[], "
                "              $8::smallint[], $9::smallint[], $10::smallint[], "
                "              $11::int[], $12::int[], $13::int[], $14::numeric[], "
                "              $15::text[], $16::text[], $17::text[]) "
                "         AS u(node_id, code, ext_code, ext_id, forma, col, niv, pos, "
                "              lx, ly, lz, peso, estado, situacion, crudo) "
                "ON CONFLICT (tenant_id, warehouse_id, external_code) "
                "  WHERE external_code IS NOT NULL AND deleted_at IS NULL "
                "DO UPDATE SET status = EXCLUDED.status, "
                "              location_situation = EXCLUDED.location_situation, "
                "              max_weight_kg = EXCLUDED.max_weight_kg, "
                "              logical_x = EXCLUDED.logical_x, "
                "              logical_y = EXCLUDED.logical_y, "
                "              logical_z = EXCLUDED.logical_z, "
                "              raw_source = EXCLUDED.raw_source, "
                "              updated_at = now()"
            )
            insertadas = 0
            lote_filas = 2000
            for i in range(0, len(filas), lote_filas):
                t = filas[i : i + lote_filas]
                await conn.execute(
                    sql_loc, tid, wid,
                    [mapa_bay[(mapa_ref[f["ref_norm"]], f["col"])] for f in t],
                    [f["code"] for f in t],
                    [f["external_code"] for f in t],
                    [f["external_location_id"] for f in t],
                    [f["forma"] for f in t],
                    [f["col"] for f in t],
                    [f["niv"] for f in t],
                    [f["pos"] for f in t],
                    [f["lx"] for f in t],
                    [f["ly"] for f in t],
                    [f["lz"] for f in t],
                    [f["peso_max"] for f in t],
                    [ESTADO_WMS.get(f["estado_wms"] or "", "available") for f in t],
                    [f["situacion"] for f in t],
                    [json.dumps({k: v for k, v in (
                        ("tipo_wms", f["tipo_wms"]), ("estado_wms", f["estado_wms"]),
                        ("zona", f["zona"]), ("preambulo", f["preambulo"]),
                        # Solo aparece cuando SE DESCARTÓ una capacidad: su
                        # presencia es la señal, igual que en 0058.
                        ("peso_max_crudo", f["peso_max_crudo"]),
                        ("capacidad_anulada_por",
                         "implausible_importador" if f["peso_max_crudo"] else None),
                    ) if v is not None}) for f in t],
                )
                insertadas += len(t)
                print(f"     … {insertadas:,} / {len(filas):,} ubicaciones", flush=True)
            print(f"  ubicaciones escritas: {insertadas:,}")

            await conn.execute(
                "UPDATE spatial.import_batches SET status='completed', "
                "  finished_at = now(), nodes_created = $2, bays_created = $3, "
                "  locations_created = $4 WHERE id = $1",
                lote, len(mapa_ref), len(mapa_bay), insertadas)

            # ── Verificación DENTRO de la transacción ───────────────────────
            n_rack = await conn.fetchval(
                "SELECT count(1) FROM spatial.nodes WHERE warehouse_id=$1 "
                "AND node_type='rack' AND deleted_at IS NULL", wid)
            n_bay = await conn.fetchval(
                "SELECT count(1) FROM spatial.nodes WHERE warehouse_id=$1 "
                "AND node_type='bay' AND deleted_at IS NULL", wid)
            n_loc = await conn.fetchval(
                "SELECT count(1) FROM spatial.locations WHERE warehouse_id=$1 "
                "AND deleted_at IS NULL", wid)
            n_world = await conn.fetchval(
                "SELECT count(1) FROM spatial.locations WHERE warehouse_id=$1 "
                "AND (world_position IS NOT NULL OR world_frame_id IS NOT NULL)", wid)
            n_aisle = await conn.fetchval(
                "SELECT count(1) FROM spatial.nodes WHERE warehouse_id=$1 "
                "AND node_type='aisle'", wid)

            print("\n  VERIFICACION en la misma transaccion")
            print(f"    racks       : {n_rack:,}")
            print(f"    cuerpos     : {n_bay:,}")
            print(f"    ubicaciones : {n_loc:,}")
            print(f"    world_* con valor : {n_world}  (0 esperado)")
            print(f"    pasillos          : {n_aisle}  (0 esperado: no se inventan)")

            if n_world:
                raise SystemExit("El importador no debe generar world_*")
            if n_aisle:
                raise SystemExit("El importador no debe inventar pasillos")

        print("\n✓ importacion completada")
        return 0
    finally:
        await conn.close()


async def main() -> int:
    ap = argparse.ArgumentParser(description="Importa el catalogo espacial del WMS")
    ap.add_argument("path")
    ap.add_argument("--warehouse-code", required=True,
                    help="El catalogo NO dice a que almacen pertenece: hay que decirlo")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true",
                    help="Ignora el corte por sha256 y reejecuta los upserts. Es la "
                         "forma de DEMOSTRAR que reimportar no duplica.")
    a = ap.parse_args()
    return await importar(
        Path(a.path), a.warehouse_code, dry_run=a.dry_run, force=a.force
    )


if __name__ == "__main__":
    # La existencia del archivo se comprueba AQUI, no dentro de la corrutina:
    # `Path.exists()` es E/S sincrona y bloquearia el bucle de eventos (ASYNC240),
    # y el bucle no aporta nada para responder «¿existe este archivo?».
    _ruta = Path(sys.argv[1]) if len(sys.argv) > 1 and not sys.argv[1].startswith("-") else None
    if _ruta is not None and not _ruta.exists():
        raise SystemExit(f"No existe {_ruta}")
    raise SystemExit(asyncio.run(main()))
