"""Importador del inventario desde ReporteInventario.xlsx.

    python tools/import_inventory_snapshot.py <ruta.xlsx> --warehouse-code OLO-CR
    python tools/import_inventory_snapshot.py <ruta.xlsx> --warehouse-code OLO-CR --dry-run

── QUÉ ES UN SNAPSHOT Y POR QUÉ NO SE «ACTUALIZA» ───────────────────────────

Un snapshot es una FOTO del inventario en un instante, y las fotos no se editan:
llega una nueva. `inventory.wms_snapshots` guarda cuándo se tomó (`taken_at`), de
dónde salió y cuántas filas traía; `inventory.wms_stock` guarda sus líneas.

Eso NO es una decisión de diseño de este importador: es la del esquema, y la razón
está en el propio dominio. El WMS es el sistema de origen y este es su espejo de solo
lectura (ADR-009 §3.4). Si aquí se pudiera «corregir» una cantidad, habría dos
verdades sobre lo que hay en un hueco, y la de este lado sería la equivocada.

Por eso importar dos veces el mismo archivo no duplica nada y tampoco mezcla: se
detecta por `sha256` antes de leer una fila.

── LO QUE ESTE IMPORTADOR **NO** HACE ───────────────────────────────────────

  · NO crea ubicaciones. Une por `location_code` contra las 29.312 que ya importó el
    catálogo espacial. Una línea cuyo hueco no existe se guarda con `location_id`
    NULO y se cuenta: es un dato real del WMS que apunta a un sitio que el catálogo
    no conoce, y ocultarlo sería esconder una discrepancia entre los dos sistemas.

  · NO crea clientes. `core.clients` tiene EPA y COFERSA dados de alta; un nombre de
    compañía que no esté se queda sin `client_id`, con el mismo criterio.

  · NO deduce ocupación. La ocupación se DERIVA de que exista una línea de stock en
    esa ubicación, y eso lo hace una vista, no el importador. Escribir un booleano
    «ocupado» aquí lo convertiría en un dato que hay que mantener sincronizado con
    las líneas que lo justifican.

── TOLERANCIA A FILAS INVÁLIDAS ────────────────────────────────────────────

Una fila mala no tumba el lote: se cuenta por motivo y el resumen los enumera. Lo que
sí tumba el lote es un error de escritura, porque entonces el snapshot quedaría a
medias y una foto a medias miente sobre lo que hay en el almacén.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import asyncpg
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

from admin_sql import _connection_kwargs

# ── Columnas del archivo ────────────────────────────────────────────────────
#
# Se nombran por su ENCABEZADO y no por su posición: el reporte tiene 77 columnas y
# el orden ha cambiado entre versiones del WMS. Buscar por nombre falla ruidosamente
# —«falta la columna X»— mientras que fiarse del índice lee el campo equivocado en
# silencio, que es peor.
COL_UBICACION = "Ubicación"
COL_ARTICULO = "Artículo"
COL_DESCRIPCION = "Descripción"
COL_CANTIDAD = "Cantidad Unidades"
COL_CANTIDAD_TXT = "Cantidad Almacenaje"
COL_PALLET = "Pallet"
COL_COMPANIA = "Nombre Compañía"
COL_LOTE = "Lote"
COL_CADUCIDAD = "Fecha Caducidad"
COL_SITUACION_UBIC = "Situación Ubicación"
COL_ESTADO_UBIC = "Estado Ubicación"

OBLIGATORIAS = (COL_UBICACION,)

# Lo que se guarda en `raw`. NO son las 77 columnas: `raw` existe para poder auditar
# una línea sin volver al Excel, no para duplicar el reporte en la base. 41.055 filas
# por 77 campos serian ~90 MB de JSONB que nadie consulta.
EN_RAW = (
    COL_ARTICULO, COL_DESCRIPCION, COL_CANTIDAD, COL_CANTIDAD_TXT, COL_PALLET,
    COL_COMPANIA, COL_LOTE, COL_CADUCIDAD, COL_SITUACION_UBIC, COL_ESTADO_UBIC,
    "Tipo Pallet", "Referencia ERP", "Código Ean", "Zona Almacenaje",
    "Tipo Ubicación", "Familia", "SubFamilia", "Peso", "Situación",
    "Fecha Ubicación", "IdAlmacenamiento", "Contenedor",
)

# Tamaño del lote de inserción. 2.000 filas por `executemany` es el punto donde el
# coste por viaje al pooler —260 ms medidos— deja de dominar sin que el paquete se
# vuelva tan grande que el servidor lo trocee.
LOTE = 2_000


def sha256_de(ruta: Path) -> str:
    h = hashlib.sha256()
    with ruta.open("rb") as f:
        for trozo in iter(lambda: f.read(1 << 20), b""):
            h.update(trozo)
    return h.hexdigest()


def _texto(v: Any) -> str | None:
    """Normaliza a texto o `None`. Un `''` y un `None` son lo mismo aquí."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _cantidad(v: Any) -> float | None:
    """La cantidad, o `None` si no la hay.

    `None` y `0` son distintos: `0` es «hay una línea y su cantidad es cero», que el
    WMS produce de verdad; `None` es «el reporte no lo dice». Convertir uno en otro
    haría que un hueco sin dato pareciera vacío.
    """
    if v is None or v == "":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return n if n >= 0 else None


def _uom(txt: Any) -> str | None:
    """La unidad de medida, de «54 UD».

    El reporte trae la cantidad dos veces: numérica en `Cantidad Unidades` y con su
    unidad pegada en `Cantidad Almacenaje`. Se toma el número de la primera y la
    unidad de la segunda, en lugar de parsear la segunda entera: si el formato
    cambiara, se perdería la unidad pero no la cantidad.
    """
    s = _texto(txt)
    if not s:
        return None
    m = re.search(r"[A-Za-z]+\s*$", s)
    return m.group(0).strip().upper()[:12] if m else None


def _fecha(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s[:19], f).date()
        except ValueError:
            continue
    return None


def _json_seguro(v: Any) -> Any:
    """Valores que `json.dumps` no sabe serializar: fechas, sobre todo."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def leer_inventario(
    ruta: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], datetime | None]:
    """Lee el archivo. Devuelve (líneas, rechazos, fecha del snapshot).

    La fecha del snapshot se toma de la `Fecha Ubicación` MÁS RECIENTE del archivo, y
    no de `now()`. Es la diferencia entre «esta foto es del almacén el martes» y «este
    archivo se subió el jueves»: con `now()`, dos importaciones del mismo reporte
    tendrían fechas distintas y no habría forma de ordenar las fotos por antigüedad
    real.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas_iter = ws.iter_rows(values_only=True)

    cab = [(_texto(c) or "") for c in next(filas_iter)]
    idx = {nombre: i for i, nombre in enumerate(cab)}
    faltan = [c for c in OBLIGATORIAS if c not in idx]
    if faltan:
        wb.close()
        raise SystemExit(
            f"El archivo no tiene la columna obligatoria {faltan!r}. "
            f"Columnas encontradas: {cab[:8]}…"
        )

    def val(fila: tuple[Any, ...], nombre: str) -> Any:
        i = idx.get(nombre)
        return None if i is None or i >= len(fila) else fila[i]

    lineas: list[dict[str, Any]] = []
    rechazos: list[dict[str, Any]] = []
    tomada: datetime | None = None

    for n, fila in enumerate(filas_iter, start=2):
        if fila is None or all(c is None or c == "" for c in fila):
            continue
        codigo = _texto(val(fila, COL_UBICACION))
        if not codigo:
            rechazos.append({"fila": n, "motivo": "sin codigo de ubicacion"})
            continue

        f_ubic = val(fila, "Fecha Ubicación")
        if isinstance(f_ubic, datetime) and (tomada is None or f_ubic > tomada):
            tomada = f_ubic

        lineas.append(
            {
                "codigo": codigo.upper(),
                "sku": _texto(val(fila, COL_ARTICULO)),
                "descripcion": _texto(val(fila, COL_DESCRIPCION)),
                "qty": _cantidad(val(fila, COL_CANTIDAD)),
                "uom": _uom(val(fila, COL_CANTIDAD_TXT)),
                "pallet": _texto(val(fila, COL_PALLET)),
                "compania": _texto(val(fila, COL_COMPANIA)),
                "lote": _texto(val(fila, COL_LOTE)),
                "caduca": _fecha(val(fila, COL_CADUCIDAD)),
                "raw": {
                    k: _json_seguro(val(fila, k))
                    for k in EN_RAW
                    if val(fila, k) not in (None, "")
                },
            }
        )

    wb.close()
    return lineas, rechazos, tomada


async def importar(ruta: Path, wh_code: str, *, dry_run: bool, force: bool) -> int:
    sha = sha256_de(ruta)
    print(f"→ archivo : {ruta.name}")
    print(f"  sha256  : {sha}")
    print(f"  almacen : {wh_code}")

    lineas, rechazos, tomada = leer_inventario(ruta)
    print(f"  lineas leidas : {len(lineas):,}   rechazadas: {len(rechazos)}")
    for m, c in Counter(r["motivo"] for r in rechazos).most_common():
        print(f"     {c:6d} x {m}")

    ubicaciones = {ln["codigo"] for ln in lineas}
    pallets = {ln["pallet"] for ln in lineas if ln["pallet"]}
    companias = Counter(ln["compania"] for ln in lineas if ln["compania"])
    print(f"  ubicaciones distintas : {len(ubicaciones):,}")
    print(f"  pallets distintos     : {len(pallets):,}")
    print(f"  companias             : {dict(companias.most_common(5))}")
    cuando = tomada.isoformat() if tomada else "(sin fecha en el archivo)"
    print(f"  foto tomada el        : {cuando}")

    if dry_run:
        print("  DRY-RUN: no se escribe nada")
        return 0

    kw = _connection_kwargs()
    print(f"→ {kw['user']}@{kw['host']}:{kw['port']}/{kw['database']}")
    conn = await asyncpg.connect(**kw)
    try:
        async with conn.transaction():
            wh = await conn.fetchrow(
                "SELECT id, tenant_id, name FROM core.warehouses "
                "WHERE code = $1 AND deleted_at IS NULL",
                wh_code,
            )
            if wh is None:
                raise SystemExit(f"No existe el almacen {wh_code}")
            tid, wid = wh["tenant_id"], wh["id"]
            print(f"  almacen resuelto: {wh['name']}")

            # ── Idempotencia · el hash del archivo ──────────────────────────
            ya = await conn.fetchval(
                "SELECT id FROM inventory.wms_snapshots "
                " WHERE tenant_id=$1 AND warehouse_id=$2 AND external_ref=$3 "
                "   AND status='ready' AND deleted_at IS NULL",
                tid, wid, sha,
            )
            if ya and not force:
                print(f"  Esta foto ya se importo (sha256 {sha[:12]}…). Nada que hacer.")
                print("  (--force la reemplaza; sirve para PROBAR la idempotencia)")
                return 0
            if ya:
                # Se borra la anterior en lugar de dejar dos fotos idénticas: dos
                # snapshots con el mismo origen y el mismo contenido no son dos
                # medidas, son la misma contada dos veces, y la ocupación saldría al
                # doble en cualquier agregado que no filtre por la última.
                borradas = await conn.fetchval(
                    "WITH d AS (DELETE FROM inventory.wms_stock WHERE snapshot_id=$1 RETURNING 1) "
                    "SELECT count(*) FROM d",
                    ya,
                )
                await conn.execute("DELETE FROM inventory.wms_snapshots WHERE id=$1", ya)
                print(f"  --force: se reemplaza la foto anterior ({borradas:,} lineas)")

            # ── El snapshot, en estado `loading` ────────────────────────────
            #
            # Se crea antes de escribir las líneas y se marca `ready` al terminar. Si
            # el proceso muere a mitad, la foto queda como `loading` y cualquier
            # consulta que pida `ready` la ignora: es la diferencia entre una foto
            # incompleta y una foto que miente.
            snap = await conn.fetchval(
                "INSERT INTO inventory.wms_snapshots "
                "(tenant_id, warehouse_id, taken_at, source, external_ref, row_count, "
                " status, notes, created_by, updated_by) "
                "VALUES ($1,$2,$3,'xlsx',$4,$5,'loading',$6, "
                "        core.current_user_id(), core.current_user_id()) "
                "RETURNING id",
                tid, wid, tomada or datetime.now().astimezone(), sha, len(lineas),
                f"Importado de {ruta.name} con tools/import_inventory_snapshot.py",
            )
            print(f"  snapshot creado: {snap}")

            # ── Ubicaciones y clientes, resueltos EN LOTE ──────────────────
            #
            # Una consulta para las 29.312 ubicaciones en vez de una por línea: 41.055
            # viajes al pooler a 260 ms serían tres horas.
            mapa_ubic = {
                r["code"]: r["id"]
                for r in await conn.fetch(
                    "SELECT upper(code) AS code, id FROM spatial.locations "
                    " WHERE warehouse_id=$1 AND deleted_at IS NULL",
                    wid,
                )
            }
            mapa_cli = {
                (r["code"] or "").upper(): r["id"]
                for r in await conn.fetch(
                    "SELECT code, id FROM core.clients WHERE tenant_id=$1 AND deleted_at IS NULL",
                    tid,
                )
            }
            # Los nombres del reporte no siempre son el código: «Cofersa» contra
            # «COFERSA». Se indexa también por nombre para no perder la relación por
            # una diferencia de mayúsculas.
            for r in await conn.fetch(
                "SELECT name, id FROM core.clients WHERE tenant_id=$1 AND deleted_at IS NULL",
                tid,
            ):
                mapa_cli.setdefault((r["name"] or "").upper(), r["id"])

            sin_ubicacion = sum(1 for ln in lineas if ln["codigo"] not in mapa_ubic)
            sin_cliente = sum(
                1 for ln in lineas
                if ln["compania"] and ln["compania"].upper() not in mapa_cli
            )
            print(f"  ubicaciones del catalogo resueltas : {len(mapa_ubic):,}")
            if sin_ubicacion:
                print(f"  ⚠ {sin_ubicacion:,} lineas apuntan a un hueco que el catalogo NO conoce")
            if sin_cliente:
                faltan = {
                    ln["compania"]
                    for ln in lineas
                    if ln["compania"] and ln["compania"].upper() not in mapa_cli
                }
                print(
                    f"  ⚠ {sin_cliente:,} lineas de companias sin dar de alta: "
                    f"{sorted(faltan)[:5]}"
                )

            # ── Las líneas ─────────────────────────────────────────────────
            registros = [
                (
                    tid, snap, wid,
                    mapa_ubic.get(ln["codigo"]),
                    ln["codigo"],
                    ln["pallet"], ln["sku"], ln["descripcion"],
                    ln["qty"], ln["uom"],
                    mapa_cli.get((ln["compania"] or "").upper()),
                    ln["lote"], ln["caduca"],
                    json.dumps(ln["raw"], ensure_ascii=False, default=str),
                )
                for ln in lineas
            ]
            escritas = 0
            for i in range(0, len(registros), LOTE):
                trozo = registros[i : i + LOTE]
                await conn.executemany(
                    "INSERT INTO inventory.wms_stock "
                    "(tenant_id, snapshot_id, warehouse_id, location_id, location_code, "
                    " pallet_code, sku, description, qty, uom, client_id, lot, expires_at, raw) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14::jsonb)",
                    trozo,
                )
                escritas += len(trozo)
                if escritas % 10_000 == 0 or escritas == len(registros):
                    print(f"    {escritas:,} / {len(registros):,}")

            await conn.execute(
                "UPDATE inventory.wms_snapshots "
                "   SET status='ready', row_count=$2, updated_by=core.current_user_id() "
                " WHERE id=$1",
                snap, escritas,
            )

            # ── Lo que quedó, contado desde la BASE ────────────────────────
            #
            # Se cuenta consultando y no confiando en el contador local: si una
            # restricción hubiera rechazado algo en silencio, el número local mentiría.
            fin = await conn.fetchrow(
                "SELECT count(*) AS lineas, "
                "       count(DISTINCT location_id) AS huecos, "
                "       count(*) FILTER (WHERE location_id IS NULL) AS huerfanas, "
                "       count(DISTINCT pallet_code) AS pallets, "
                "       sum(qty) AS unidades "
                "  FROM inventory.wms_stock WHERE snapshot_id=$1",
                snap,
            )
            print("\n  ── EN LA BASE ──")
            print(f"  lineas          : {fin['lineas']:,}")
            print(f"  huecos ocupados : {fin['huecos']:,}")
            print(f"  sin hueco       : {fin['huerfanas']:,}")
            print(f"  pallets         : {fin['pallets']:,}")
            print(
                f"  unidades        : {fin['unidades']:,}"
                if fin["unidades"]
                else "  unidades        : -"
            )
    finally:
        await conn.close()
    print("\n✓ OK")
    return 0


async def main() -> int:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("archivo", type=Path)
    p.add_argument("--warehouse-code", required=True)
    p.add_argument("--dry-run", action="store_true", help="lee y resume, sin escribir")
    p.add_argument(
        "--force",
        action="store_true",
        help="reemplaza una foto ya importada con el mismo sha256",
    )
    a = p.parse_args()
    if not a.archivo.exists():
        raise SystemExit(f"No existe el archivo {a.archivo}")
    return await importar(a.archivo, a.warehouse_code, dry_run=a.dry_run, force=a.force)


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
